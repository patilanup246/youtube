# -*- coding: utf-8 -*-
import scrapy
import MySQLdb
import openpyxl
from scrapy.crawler import CrawlerProcess
import sys


class AllabolaSpider(scrapy.Spider):
    name = 'allabola'
    allowed_domains = ['https://www.allabolag.se']
    start_urls = []
    #'https://www.allabolag.se/7696250484/befattningar'
    host = '104.197.180.57'
    user = 'root'
    password = 'root'
    DB_name = "db_allabolag"
    f = open('Facebook_Auidance.csv', 'w')
    f.write('fn,ln,zip,ct,st,country,dob,doby,gen,age,uid')
    f.write('\n')
    f.close()
    try:
        connection = MySQLdb.connect(host, user, password,DB_name ,charset='utf8')
        cursor = connection.cursor()
    except Exception as e:
        print(str(e))

    try:
        strquery2 = "CREATE TABLE tbl_allabola""""(Id INT NOT NULL AUTO_INCREMENT,
                                                        Registration_no varchar(250) DEFAULT NULL,
                                                        First_name varchar(250) DEFAULT NULL,
                                                        Middle_name varchar(250) DEFAULT NULL,
                                                        Famaily_name varchar(250) DEFAULT NULL,
                                                        Gender longtext DEFAULT NULL,
                                                        Year longtext DEFAULT NULL,
                                                        Board_member longtext DEFAULT NULL,
                                                        PRIMARY KEY (`Id`))"""

        cursor.execute(strquery2)
    except Exception as e:
        print(str(e))

    def start_requests(self):
        try:

            wb = openpyxl.load_workbook(
                '/home//Business_numbers.xlsx')
            ws = wb.get_active_sheet()

            row_count = ws.max_row



            for h in range(2,row_count):
                regi_number = ws.cell(row=h, column=2).value
                Post_Code = ws.cell(row=h, column=4).value
                main_link  = 'https://www.allabolag.se/'+str(regi_number)+'/befattningar'
                yield scrapy.FormRequest(main_link,callback=self.parse,dont_filter=True,meta={'Post_Code':Post_Code})
        except Exception as e:
            print(e)

    def parse(self, response):

        Post_Code = response.meta['Post_Code']
        Registration_no = response.url
        Registration_no = Registration_no.split('.se/')[1]
        Registration_no = Registration_no.split('/')[0]
        print(Registration_no)
        ALl_data = response.xpath('//*[@class="list--personnel accordion-body"]/li')

        for datas in ALl_data:

            gender = datas.xpath(".//div[1]/span[contains(@class,'male')]/@class").extract_first()
            gender = gender.split('--')[1]
            gender = gender.encode('utf-8')
            if gender == 'male':
                gender = 'm'
            elif gender == 'female':
                gender = 'f'

            name = datas.xpath('.//div[2]/a/text()').extract_first()
            name = name.strip()
            name = name.split(' (f. ')
            year = name[1].replace(')','')
            if year != None:
                age = str(2019 - int(year))
            fullname = name[0]
            # try:
            #     fullname = str(fullname)
            # except Exception as e:
            #     print e
            fullname = fullname.split(' ')
            firstname = ''
            middlename = ''
            familyname = ''
            if len(fullname) == 3:
                firstname = fullname[0]
                middlename = fullname[1]
                familyname = fullname[2]
            elif len(fullname) == 2:
                firstname = fullname[0]
                middlename = fullname[1]
            elif len(fullname) > 3:
                firstname = fullname[0]
                familyname = fullname[-1]
                middlename = ''
                for k in range(1,len(fullname)-1):
                    if middlename == '':
                        middlename = fullname[k]
                    else:
                        middlename = middlename + ' ' + fullname[k]


            type = datas.xpath('.//div[2]/text()').extract()[2]
            Board_member = type.replace('\n','').strip()
            if gender != '':

                f = open('Facebook_Auidance.csv', 'a')
                try:
                    f.write(firstname+','+familyname+','+Post_Code+','+''+','+''+','+'Sweden'+','+''+','+year+','+gender+','+age+','+'')
                except Exception as e:
                    f.close()
                try:
                    f.write('\n')
                    f.close()
                except Exception as e:
                    ''

            if gender != '':
                try:
                    reload(sys)
                    sys.setdefaultencoding('utf8')
                    self.cursor.execute(
                        """INSERT INTO tbl_allabola(Registration_no,First_name,Middle_name,Famaily_name,Gender,Year,Board_member)VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                        (Registration_no, firstname, middlename,familyname,gender,year,Board_member))
                    self.connection.commit()
                except Exception as e:
                    print(e)


process = CrawlerProcess({'LOG_ENABLED': False})
process.crawl(AllabolaSpider)
try:
    process.start()
except:
    pass


